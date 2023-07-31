import openpyxl as xl
from openpyxl.chart import bar_chart,reference
wb=xl.load_workbook('transaction.xlsx')
 sheet=wb['Sheet1']
cell= sheet['a1']
cell=sheet.cell(1,1)

for row in range (2,sheet.max_row +1):
 cell=sheet.cell(row,3)
 corrected_price=cell.value*0.9
 corrected_price_cell=sheet.cell(row,4)
 corrected_price.value=corrected_price_cell
 
 values =reference(sheet,
           min_row=2,
           max_row=sheet.max_row,
           min_col=4,
           max_col=4)
 chart=bar_chart
 chart.add_data(values)
 sheet.add_chart(chart,'e2')
wb.save('transaction2.xlsx')